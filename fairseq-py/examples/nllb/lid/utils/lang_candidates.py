#!/usr/bin/env python3 -u

import argparse
import math

import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter


# fmt: off
lid_187_codes = ["af", "als", "am", "an", "ar", "arz", "as", "ast", "av", "az", "azb", "ba", "bar", "bcl", "be", "bg", "bh", "bn", "bo", "bpy", "br", "bs", "bxr", "ca", "cbk", "ce", "ceb", "ckb", "co", "cs", "cv", "cy", "da", "de", "diq", "dsb", "dty", "dv", "el", "eml", "en", "eo", "es", "et", "eu", "fa", "fi", "fr", "frr", "fy", "ga", "gd", "gl", "gn", "gom", "gu", "gv", "ha", "he", "hi", "hif", "hr", "hsb", "ht", "hu", "hy", "ia", "id", "ie", "ig", "ilo", "io", "is", "it", "ja", "jbo", "jv", "ka", "kab", "kk", "km", "kn", "ko", "krc", "ku", "kv", "kw", "ky", "la", "lb", "lez", "lg", "li", "lmo", "ln", "lo", "lrc", "lt", "lv", "mai", "mg", "mhr", "min", "mk", "ml", "mn", "mr", "mrj", "ms", "mt", "mwl", "my", "myv", "mzn", "nah", "nap", "nds", "ne", "new", "nl", "nn", "no", "oc", "om", "or", "os", "pa", "pam", "pfl", "pl", "pms", "pnb", "ps", "pt", "qu", "rm", "ro", "ru", "rue", "sa", "sah", "sat", "sc", "scn", "sco", "sd", "si", "sk", "sl", "sn", "so", "sq", "sr", "su", "sv", "sw", "ta", "te", "tg", "th", "tk", "tl", "tn", "tr", "tt", "tyv", "ug", "uk", "ur", "uz", "vec", "vep", "vi", "vls", "vo", "wa", "war", "wo", "wuu", "xal", "xh", "xmf", "yi", "yo", "yue", "zh", "zu"]
flores_124_codes = ["afr", "amh", "ara", "ara-IQ", "ara-LB", "ara-MA", "ara-SA", "ara-TN", "ara-YE", "asm", "ast", "aym", "azj", "bel", "ben", "bos", "bul", "cat", "ceb", "ces", "ckb", "cym", "dan", "deu", "dyu", "ell", "eng", "est", "fas", "fin", "fra", "ful", "gla", "gle", "glg", "guj", "hat", "hau", "heb", "hin", "hrv", "hun", "hye", "ibo", "ind", "isl", "ita", "jav", "jpn", "kac", "kam", "kan", "kat", "kaz", "kea", "khm", "kir", "kmb", "kon", "kor", "kur", "lao", "lav", "lin", "lit", "ltz", "lug", "luo", "mal", "mar", "mkd", "mlg", "mlt", "mon", "mri", "msa", "mya", "nld", "nob", "npi", "nso", "nya", "oci", "orm", "ory", "pan", "pol", "por", "pus", "que", "ron", "rus", "sin", "slk", "slv", "sna", "snd", "som", "spa", "sqi", "srp", "ssw", "sun", "swe", "swh", "tam", "tel", "tgk", "tgl", "tha", "tir", "tsn", "tur", "ukr", "umb", "urd", "uzb", "vie", "wol", "xho", "yid", "yor", "yue", "zho_simpl", "zho_trad", "zul"]
mid_mine_codes = ["cym", "gla", "gle", "mya", "kac", "hat", "kea", "mlt", "kaz", "kir", "uzb", "ckb", "kur", "tgk", "ceb", "jav", "mri", "sun", "aym", "que"]


iso_639_3_to_flores_124 = {
    "nor": "nob",
    "zho": "zho_simpl"
}

# fmt: on


lid_uniq_id_col_name = "ISO 639-3"  # to be replaced by BCP-47?
has_lid_col_name = "has lid data"


class InferFromRow:
    def __init__(self, name, codes, iso_639_3_mapper={}):
        self.name = name
        self.seen_codes = set()
        self.codes = set(codes)
        self.iso_639_3_mapper = iso_639_3_mapper

    def __call__(self, row):
        wiki_code = row["Wiki code"]
        iso_639_3_code = row[lid_uniq_id_col_name]

        return self.infer_from_codes(wiki_code, iso_639_3_code)

    def infer_from_codes(self, wiki_code, iso_639_3_code):
        val = None
        if wiki_code in self.codes:
            val = wiki_code
        if iso_639_3_code in self.codes:
            if val and val != iso_639_3_code:
                print(f"{self.name}: same code? {wiki_code} {iso_639_3_code}")
            val = iso_639_3_code
        if self.iso_639_3_mapper.get(iso_639_3_code, None) in self.codes:
            assert val == "" or val is None
            val = self.iso_639_3_mapper[iso_639_3_code]
        if val:
            if val in self.seen_codes:
                print(f"Warning: {val} already added")
            self.seen_codes.add(val)

        return val

    def check_unseen_codes(self):
        unseen = sorted(list(self.codes.difference(self.seen_codes)))
        print(self.name, f"{len(unseen)} unseen:\n\t", "\n\t".join(unseen))


class InferLID187(InferFromRow):
    def __init__(self):
        super().__init__("lid_187", lid_187_codes)


class InferFlores124(InferFromRow):
    def __init__(self):
        super().__init__("flores_124", flores_124_codes, iso_639_3_to_flores_124)


class InferMidMine(InferFromRow):
    def __init__(self):
        super().__init__("mid_mine", mid_mine_codes)


def fill_df(df, inference_classes):
    print("")
    print("Inferring new columns")
    for inference_class in inference_classes:
        inference = inference_class()
        values = []
        indices = []
        for index, row in df.iterrows():
            val = inference(row)
            values.append(val)
            indices.append(index)

        df.loc[indices, f"{inference.name} code"] = values
        inference.check_unseen_codes()

    print("")


def fill_has_lid_data(df):
    has_lid_data_cond = (df["JW300 code"].notna()) | (df["lid187 code"].notna())
    has_lid_data = df[has_lid_data_cond]
    df[has_lid_col_name] = np.where(has_lid_data_cond, "Yes", "")


def detect_remove_duplicates(df):
    vals = df[lid_uniq_id_col_name]
    if vals.isna().any():
        print(f"Has NaN value(s): {df[vals.isna()]}")
    vals, count = np.unique(vals.dropna().values, return_counts=True)
    duplicates = vals[count > 1]
    print(f"duplicates = {duplicates}")

    for dpl in duplicates:
        dpl_rows = df[df[lid_uniq_id_col_name] == dpl]
        df.drop(
            dpl_rows[["JW300 code", "lid187 code"]]
            .isna()
            .sum(axis=1)
            .sort_values()[1:]
            .index,
            inplace=True,
        )


def print_seed_data_collection(df):
    print("")
    print("Seed data Collection")
    print(
        df[df[has_lid_col_name] != "Yes"][
            ["English name", lid_uniq_id_col_name, "Seed Data"]
        ].sort_values(by=["Seed Data"])
    )
    print("\n")


def check_consistency_after_merge(df, columns_to_check, right_suffix):
    for col in columns_to_check:
        if (df[col] != df[col + right_suffix]).any():
            print(f'After merge of two sheets, column "{col}" is different for:')
            print(
                df[df[col] != df[col + right_suffix]][
                    [lid_uniq_id_col_name, col, col + right_suffix]
                ]
            )


def get_summary_sheet(filename):
    df = pd.read_excel(
        filename,
        sheet_name="Summary Sheet",
        header=8,
    )

    df = df.rename(
        columns={
            "Unnamed: 1": "NLLB_OR_NOT",
            "Unnamed: 2": "Annotation",
            "Unnamed: 3": "Annotation Notes",
        }
    )

    df = df[df["NLLB_OR_NOT"] == "NLLB-2021"]
    df.dropna(subset=["English name"], inplace=True)

    speakr_q = np.quantile(df['# Speakers'].dropna().values, q=0.10)
    articl_q = np.quantile(df['# Articles'].dropna().values, q=0.30)

    low_resource_df = df[(df['# Speakers'] < speakr_q) | (df['# Articles'] < articl_q)]

    low_resource_langs = ['bam', 'epo', 'kin', 'ady', 'aka', 'awa', 'bjn', 'bis', 'che',
       'chr', 'nya', 'din', 'dzo', 'ewe', 'fij', 'fon', 'gom', 'kal',
       'grn', 'haw', 'kbp', 'kau', 'krc', 'kas', 'kik', 'kon', 'ltg',
       'mni', 'nia', 'pag', 'pap', 'roh', 'run', 'bxr', 'smo', 'sag',
       'skr', 'alt', 'sot', 'tah', 'bod', 'tpi', 'tog', 'tso', 'tum',
       'twi', 'uig', 'cre', 'iku', 'aym', 'bos', 'est', 'ful', 'lug',
       'ibo', 'gle', 'kmb', 'lao', 'lav', 'lin', 'lit', 'mkd', 'mlt',
       'ary', 'orm', 'slv', 'ssw', 'tir', 'tsn', 'wol', 'xho']

    columns_to_be_kept = [
        "English name",
        lid_uniq_id_col_name,
        "NLLB_OR_NOT",
        "Annotation",
        "Annotation Notes",
        "Script",
        "JW300 code",
        "lid187 code",
        "Wiki code",
    ]
    df = df[columns_to_be_kept].copy()

    return df


def get_status_q3_sheet(filename):
    headers = [
        "English name",
        "African?",
        lid_uniq_id_col_name,
        "LID, fasttext precision",
        "FLORES Status",
        "Seed Data",
        "JW300 Code",
        "Human Eval",
        "UXR Study",
        "Wikipedia Launch",
        "Monolingual Data",
        "Bilingual Data",
        "Script",
        "Bilingual Data 2",
    ]
    df_q3 = pd.read_excel(filename, sheet_name="Status Q3", header=2, names=headers)

    # fix google docs export issue
    b_col1 = "Bilingual Data"
    b_col2 = "Bilingual Data 2"
    df_q3[b_col1] = np.where(df_q3[b_col1].isna(), df_q3[b_col2], df_q3[b_col1])

    columns_to_be_kept_q3 = [
        "English name",
        lid_uniq_id_col_name,
        "African?",
        "FLORES Status",
        "Seed Data",
        "Bilingual Data",
        "Script",
    ]
    df_q3 = df_q3[columns_to_be_kept_q3].copy()

    return df_q3


def save_file(df, out_filename, nb_columns_before_inference):
    with pd.ExcelWriter(out_filename) as writer:
        df.to_excel(writer, sheet_name="Languages")
        wb = writer.book
        ws = wb.active
        c = ws["C2"]
        ws.freeze_panes = c
        ws.insert_cols(nb_columns_before_inference + 2)
        ws.column_dimensions[
            get_column_letter(nb_columns_before_inference + 2)
        ].width = 30

        wb.save(out_filename)


def export_bash_statements(df):
    goal124_df = df[(df[has_lid_col_name] == "Yes") | ~df["flores_124 code"].isna()].copy()
    goal124_df.fillna("",inplace=True)
    export_bash_sources = [
        ("ISO_639_3_LANGS", "ISO 639-3"),
        ("JW300_LANGS", "JW300 code"),
        ("LID_187_LANGS", "lid187 code"),
        ("FLORES_LANGS", "flores_124 code"),
    ]

    export_dfs = [
        ("GOAL124", goal124_df),
    ]

    for export_df_name, bash_export_df in export_dfs:
        print(f"\nBash exports {export_df_name} (size: {len(bash_export_df)}):\n")
        for bash_var_name, column_name in export_bash_sources:
            vals = " ".join([f'"{val}"' for val in bash_export_df[column_name].values])
            print(f"{export_df_name}__{bash_var_name}=({vals})")
        print("")


def main():
    parser = argparse.ArgumentParser(
        description="This script takes the NLLB Language candidates spreadsheet as input and adds some columns useful for LID."
    )
    parser.add_argument(
        "--input",
        default="NLLB Language Candidates.xlsx",
        type=str,
        help="Exported NLLB Language Candidate excel file",
    )
    parser.add_argument(
        "--output",
        default="mappings.xlsx",
        type=str,
        help="Output Excel file that contains the mappings",
    )
    args = parser.parse_args()

    df = get_summary_sheet(args.input)
    df_q3 = get_status_q3_sheet(args.input)


    outer_df = df.merge(df_q3, on=lid_uniq_id_col_name, suffixes=["", "_q3"], how="outer")
    left_df = df.merge(df_q3, on=lid_uniq_id_col_name, suffixes=["", "_q3"], how="left")
    outer_set = set(outer_df["ISO 639-3"].values.tolist())
    left_set = set(left_df["ISO 639-3"].values.tolist())
    diff_list = '\n'.join(sorted(list(outer_set.difference(left_set))))
    if diff_list:
        print(f"/!\\ missing from outer join:")
        print(diff_list)

    out_df = left_df

    check_consistency_after_merge(out_df, ["Script", "English name"], "_q3")
    nb_columns_before_inference = len(out_df.columns)

    fill_has_lid_data(out_df)
    fill_df(out_df, [InferFlores124, InferMidMine])
    out_df = out_df.sort_values(by=["mid_mine code", "flores_124 code"])
    detect_remove_duplicates(out_df)
    print_seed_data_collection(out_df)

    export_bash_statements(out_df)
    save_file(out_df, args.output, nb_columns_before_inference)


if __name__ == "__main__":
    main()
